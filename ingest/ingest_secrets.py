#!/usr/bin/env python3
"""
UNSEALED — record ingester.

Sources are configured in ingest/sources.json. For each source group the
script will:
  1. Fetch the listing URL(s) (with TTL-aware on-disk cache)
  2. Discover linked release artifacts (.xlsx / .csv / detail HTML / API)
  3. Parse rows into the canonical schema
  4. Emit SQL INSERT OR IGNORE statements to db/ingest_<group>.sql
  5. Emit a health report to db/ingest_report.json
  6. Track newly-seen artifacts in ingest/seen_artifacts.json (hub-diff)

Output is intentionally SQL files (not direct D1 writes) so they can be
applied with `wrangler d1 execute unsealed --remote --file=...` and stay
inspectable / reversible. INSERT OR IGNORE on the content_hash column
makes re-applying the same SQL safe.

Usage:
  python3 ingest/ingest_secrets.py              # ingest all groups
  python3 ingest/ingest_secrets.py --group nara_ndc
  python3 ingest/ingest_secrets.py --dry-run    # parse, don't write SQL
  python3 ingest/ingest_secrets.py --limit 1000 # cap rows per group
  python3 ingest/ingest_secrets.py --no-cache   # bypass on-disk cache
  python3 ingest/ingest_secrets.py --cache-ttl 0  # ignore cache age

Dependencies (install once):
  pip3 install requests beautifulsoup4 openpyxl curl_cffi
  (curl_cffi is only required for sources with "impersonate" set in sources.json,
  currently aaro.mil — Akamai blocks vanilla urllib regardless of headers.)
"""
from __future__ import annotations

import argparse
import calendar
import datetime as _dt
import hashlib
import json
import re
import sys
import time
from dataclasses import dataclass, asdict, field
from pathlib import Path
from urllib.parse import urljoin, urlparse

try:
    import requests
    from bs4 import BeautifulSoup
except ImportError:
    sys.stderr.write(
        "Missing deps. Run:  pip3 install requests beautifulsoup4 openpyxl\n"
    )
    sys.exit(1)

# Optional: curl_cffi gives us real-browser TLS fingerprints, which is the
# only way past Akamai-fronted .mil sources (e.g. aaro.mil) that JA3-block
# vanilla requests / urllib regardless of headers. Loaded lazily so the
# ingester still works without it for groups that don't need it.
try:
    from curl_cffi import requests as _cffi_requests  # type: ignore
    _HAS_CFFI = True
except ImportError:
    _HAS_CFFI = False

ROOT = Path(__file__).resolve().parent.parent
SOURCES_FILE = ROOT / "ingest" / "sources.json"
OUT_DIR = ROOT / "db"
CACHE_DIR = ROOT / "ingest" / "cache"
CACHE_DIR.mkdir(parents=True, exist_ok=True)
SEEN_ARTIFACTS_FILE = ROOT / "ingest" / "seen_artifacts.json"
SEEN_HASHES_FILE = ROOT / "ingest" / "seen_hashes.json"

USER_AGENT = (
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
    "AppleWebKit/605.1.15 (KHTML, like Gecko) "
    "Version/17.4 Safari/605.1.15"
)
FROM_HEADER = "unsealed-bot@github.com/mozltovcoktail/Unsealed"
TIMEOUT = 30

# Default cache TTL: hub pages get 7 days (we want to see new releases),
# content-addressed artifacts (dated .xlsx) effectively never expire.
DEFAULT_HUB_TTL_SEC = 7 * 24 * 3600
DEFAULT_ARTIFACT_TTL_SEC = 365 * 24 * 3600

RUN_ID = _dt.datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")


# ─── Canonical record ─────────────────────────────────────────────
@dataclass
class Record:
    title: str
    agency: str
    unsealed_date: str  # ISO 8601
    collection_id: str | None
    source_url: str
    description: str | None = None
    thumbnail_url: str | None = None
    source_artifact_url: str | None = None  # provenance
    is_sealed: int = 0  # 1 = record is on a candidate / pre-release list, not actually unsealed
    document_date: str | None = None  # doc's own creation date (FRUS: doc date; unsealed_date = volume pub date)
    content_hash: str = ""

    def __post_init__(self):
        if not self.content_hash:
            h = hashlib.sha1()
            h.update((self.agency or "").encode("utf-8"))
            h.update(b"|")
            h.update((self.title or "").encode("utf-8"))
            h.update(b"|")
            h.update((self.source_url or "").encode("utf-8"))
            self.content_hash = h.hexdigest()


# ─── HTTP with on-disk cache ──────────────────────────────────────
def _cache_paths(url: str, binary: bool) -> tuple[Path, Path]:
    key = re.sub(r"[^a-zA-Z0-9]+", "_", url)[:160]
    suffix = ".bin" if binary else ".txt"
    return CACHE_DIR / f"{key}{suffix}", CACHE_DIR / f"{key}.meta.json"


_LAST_FETCH_AT: dict[str, float] = {}


def fetch(
    url: str,
    *,
    binary: bool = False,
    no_cache: bool = False,
    ttl_sec: int | None = None,
    extra_headers: dict | None = None,
    impersonate: str | None = None,
    crawl_delay_sec: int = 0,
) -> bytes | str:
    cache, meta = _cache_paths(url, binary)
    if not no_cache and cache.exists():
        age = time.time() - cache.stat().st_mtime
        max_age = ttl_sec if ttl_sec is not None else DEFAULT_HUB_TTL_SEC
        if max_age <= 0 or age < max_age:
            return cache.read_bytes() if binary else cache.read_text("utf-8", "replace")
    # Polite rate-limit: honor per-host crawl-delay between non-cached requests.
    if crawl_delay_sec > 0:
        host = urlparse(url).netloc
        last = _LAST_FETCH_AT.get(host, 0.0)
        wait = crawl_delay_sec - (time.time() - last)
        if wait > 0:
            time.sleep(wait)
        _LAST_FETCH_AT[host] = time.time()
    print(f"  fetch {url}", file=sys.stderr)
    headers = {"user-agent": USER_AGENT, "From": FROM_HEADER}
    if extra_headers:
        headers.update(extra_headers)
    if impersonate:
        if not _HAS_CFFI:
            raise RuntimeError(
                f"source needs impersonate={impersonate!r} but curl_cffi is not installed. "
                "Run: pip3 install curl_cffi"
            )
        # curl_cffi sets its own UA/headers based on the impersonate profile.
        # We pass extra_headers (e.g. API keys) but not our default UA, which
        # would override the profile's fingerprint-consistent set.
        cffi_headers = dict(extra_headers) if extra_headers else None
        r = _cffi_requests.get(url, headers=cffi_headers, impersonate=impersonate, timeout=TIMEOUT)
    else:
        r = requests.get(url, headers=headers, timeout=TIMEOUT)
    r.raise_for_status()
    if binary:
        cache.write_bytes(r.content)
    else:
        cache.write_text(r.text, "utf-8")
    meta.write_text(
        json.dumps(
            {
                "url": url,
                "fetched_at": _dt.datetime.utcnow().isoformat() + "Z",
                "etag": r.headers.get("ETag"),
                "last_modified": r.headers.get("Last-Modified"),
                "status": r.status_code,
            }
        ),
        "utf-8",
    )
    return r.content if binary else r.text


# ─── Hub-diff: track artifact URLs across runs ────────────────────
def load_seen_artifacts() -> dict[str, list[str]]:
    if SEEN_ARTIFACTS_FILE.exists():
        try:
            return json.loads(SEEN_ARTIFACTS_FILE.read_text("utf-8"))
        except Exception:
            pass
    return {}


def save_seen_artifacts(seen: dict[str, list[str]]) -> None:
    SEEN_ARTIFACTS_FILE.write_text(json.dumps(seen, indent=2, sort_keys=True), "utf-8")


def load_seen_hashes() -> set[str]:
    """Content-hashes already in remote D1. Used to skip re-emitting rows that
    would otherwise INSERT OR IGNORE — those still count against D1's
    100k-writes/day free-tier quota even when they're no-ops.

    Populated by the workflow via:
      wrangler d1 execute unsealed --remote --json \\
        --command 'SELECT content_hash FROM records' > ingest/seen_hashes.json
    """
    if not SEEN_HASHES_FILE.exists():
        return set()
    try:
        raw = json.loads(SEEN_HASHES_FILE.read_text("utf-8"))
    except Exception:
        return set()
    # Accept several shapes:
    #   ["sha1", "sha1", ...]                                  bare list
    #   [{"content_hash": "sha1"}, ...]                        rows directly
    #   [{"results": [{"content_hash": "sha1"}, ...], ...}]    wrangler --json
    #   {"results": [...]}                                     single envelope
    def _extract_rows(node):
        if isinstance(node, list):
            return node
        if isinstance(node, dict) and isinstance(node.get("results"), list):
            return node["results"]
        return []

    rows = _extract_rows(raw)
    if rows and isinstance(rows[0], dict) and "results" in rows[0]:
        # Wrangler wraps the actual rows one level deeper.
        rows = _extract_rows(rows[0])
    out: set[str] = set()
    for r in rows:
        if isinstance(r, str):
            out.add(r)
        elif isinstance(r, dict):
            h = r.get("content_hash")
            if h:
                out.add(h)
    return out


# ─── NARA NDC parser ──────────────────────────────────────────────
def parse_nara_ndc(group_cfg: dict, limit: int | None, *, fetch_opts) -> tuple[list[Record], list[str]]:
    out: list[Record] = []
    seen_xlsx: set[str] = set()

    for hub in group_cfg["urls"]:
        html = fetch(hub, **fetch_opts(is_hub=True))
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            href = urljoin(hub, a["href"])
            if href.lower().endswith((".xlsx", ".xls", ".csv")):
                seen_xlsx.add(href)

    print(f"[nara_ndc] {len(seen_xlsx)} release artifacts discovered", file=sys.stderr)
    artifacts = sorted(seen_xlsx)
    for art in artifacts:
        try:
            rows = _parse_release_artifact(art, fetch_opts=fetch_opts)
        except Exception as e:
            print(f"  skip {art}: {e}", file=sys.stderr)
            continue
        # IOD-candidate spreadsheets list entries proposed for declassification,
        # not yet released. Tag them so the UI can filter them out by default.
        sealed = 1 if "iod-candidate" in art.lower() else 0
        for row in rows:
            out.append(
                Record(
                    title=row.get("title", "").strip() or "(untitled)",
                    agency=group_cfg["agency"],
                    unsealed_date=row.get("unsealed_date") or "",
                    collection_id=row.get("collection_id")
                    or group_cfg["collection_id_default"],
                    source_url=row.get("source_url") or art,
                    description=row.get("description"),
                    source_artifact_url=art,
                    is_sealed=sealed,
                )
            )
            if limit and len(out) >= limit:
                return out, artifacts
    return out, artifacts


def _parse_release_artifact(url: str, *, fetch_opts) -> list[dict]:
    if url.lower().endswith(".csv"):
        text = fetch(url, **fetch_opts(is_hub=False))
        return _parse_csv_text(text, source_url=url)
    from openpyxl import load_workbook
    import io

    data = fetch(url, binary=True, **fetch_opts(is_hub=False))
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    iso_date = _date_from_filename(url)
    rows: list[dict] = []
    for sheet in wb.worksheets:
        header = None
        for r in sheet.iter_rows(values_only=True):
            if header is None:
                if r and any(
                    isinstance(c, str) and ("title" in c.lower() or "entry" in c.lower())
                    for c in r if c
                ):
                    header = [
                        (c or "").strip().lower() if isinstance(c, str) else ""
                        for c in r
                    ]
                continue
            row = dict(zip(header, r))
            mapped = _map_ndc_row(row, default_date=iso_date, source_url=url)
            if mapped:
                rows.append(mapped)
    return rows


_FY_RX = re.compile(r"(?P<q>1st|2nd|3rd|4th)[^a-z0-9]+quarter[^a-z0-9]+(?:release[^a-z0-9]+)?(?:list[^a-z0-9]+)?(?:fy[^a-z0-9]*)?(?P<y>\d{2,4})", re.I)
_YEAR_FIRST_RX = re.compile(r"(?P<y>\d{4})[^a-z0-9]+ndc[^a-z0-9]+(?P<q>1st|2nd|3rd|4th)", re.I)
# FY2019-Q2, FY2020_Q1, fy2024q3 etc.
_FY_COMPACT_RX = re.compile(r"fy[-_]?(?P<y>20\d\d)[-_]?q(?P<q>[1-4])", re.I)
# 2023-3rd-quarter, 2024_4th_quarter
_CY_ORDINAL_RX = re.compile(
    r"(?P<y>20\d\d)[-_](?P<q>[1-4])(?:st|nd|rd|th)[-_]?quarter", re.I
)
# q4-2022, q1-2024 (quarter then year)
_CY_QYYYY_RX = re.compile(r"q(?P<q>[1-4])[-_](?P<y>20\d\d)", re.I)
# q1-february-23 (quarter, month-word, 2-digit year)
_CY_QMONTH_YY_RX = re.compile(
    r"q(?P<q>[1-4])[-_](?:jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*[-_](?P<y>\d{2})\b",
    re.I,
)
# 2nd-qt-2023 (ordinal, qt/qtr, year)
_CY_ORD_QT_RX = re.compile(
    r"(?P<q>[1-4])(?:st|nd|rd|th)[-_](?:qt|qtr)[-_](?P<y>20\d\d)", re.I
)
# released-entries-05-12.xls → May 2012 (month-end)
_RELEASED_MONTH_RX = re.compile(
    r"released-entries-(?P<mo>\d{2})-(?P<y>\d{2})\.xlsx?$", re.I
)


def _q_to_iso(q: int, y: int, *, fiscal: bool) -> str:
    if fiscal:
        return {1: f"{y - 1:04d}-12-31", 2: f"{y:04d}-03-31",
                3: f"{y:04d}-06-30", 4: f"{y:04d}-09-30"}[q]
    return {1: f"{y:04d}-03-31", 2: f"{y:04d}-06-30",
            3: f"{y:04d}-09-30", 4: f"{y:04d}-12-31"}[q]


def _date_from_filename(url: str) -> str:
    name = url.rsplit("/", 1)[-1].lower()
    m = _RELEASED_MONTH_RX.search(name)
    if m:
        mo, y = int(m.group("mo")), 2000 + int(m.group("y"))
        if 1 <= mo <= 12:
            last = calendar.monthrange(y, mo)[1]
            return f"{y:04d}-{mo:02d}-{last:02d}"
    m = _FY_COMPACT_RX.search(name)
    if m:
        return _q_to_iso(int(m.group("q")), int(m.group("y")), fiscal=True)
    m = _CY_ORDINAL_RX.search(name)
    if m:
        return _q_to_iso(int(m.group("q")), int(m.group("y")), fiscal=False)
    m = _CY_ORD_QT_RX.search(name)
    if m:
        return _q_to_iso(int(m.group("q")), int(m.group("y")), fiscal=False)
    m = _CY_QMONTH_YY_RX.search(name)
    if m:
        return _q_to_iso(int(m.group("q")), 2000 + int(m.group("y")), fiscal=False)
    m = _CY_QYYYY_RX.search(name)
    if m:
        return _q_to_iso(int(m.group("q")), int(m.group("y")), fiscal=False)
    for rx in (_YEAR_FIRST_RX, _FY_RX):
        m = rx.search(name)
        if not m:
            continue
        q = {"1st": 1, "2nd": 2, "3rd": 3, "4th": 4}[m.group("q").lower()]
        y = int(m.group("y"))
        if y < 100:
            y += 2000
        return _q_to_iso(q, y, fiscal="fy" in name)
    return ""


def _parse_csv_text(text: str, source_url: str = "") -> list[dict]:
    import csv, io as _io

    rdr = csv.DictReader(_io.StringIO(text))
    iso = _date_from_filename(source_url) if source_url else ""
    return [
        m
        for m in (
            _map_ndc_row({k.lower(): v for k, v in r.items()}, default_date=iso, source_url=source_url)
            for r in rdr
        )
        if m
    ]


def _map_ndc_row(row: dict, default_date: str = "", source_url: str = "") -> dict | None:
    title = (
        row.get("record entry title")
        or row.get("entry title")
        or row.get("title")
        or row.get("series title")
        or row.get("collection title")
        or ""
    )
    if not title:
        return None
    def _s(v):
        return "" if v is None else str(v).strip()

    rg = _s(row.get("rg") or row.get("record group"))
    office = _s(row.get("office"))
    custodial = _s(row.get("custodial unit"))
    media = _s(row.get("media type"))
    hms_id = _s(
        row.get("hms record entry id# ")
        or row.get("hms record entry id#")
        or row.get("hms entry")
    )

    desc_bits = [b for b in (office, custodial, media) if b]
    description = " // ".join(desc_bits) if desc_bits else None

    date = (
        row.get("release date")
        or row.get("date released")
        or row.get("declass date")
        or ""
    )
    iso = _to_iso_date(str(date)) or default_date

    return {
        "title": str(title).strip(),
        "collection_id": (f"RG {str(rg).strip()}" if rg else None),
        "unsealed_date": iso,
        "description": description,
        "source_url": (f"{source_url}#hms-{hms_id}" if hms_id and source_url else source_url) or None,
    }


_DATE_RX = [
    re.compile(r"^(\d{4})-(\d{2})-(\d{2})"),
    re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})"),
    re.compile(r"^(\d{4})/(\d{1,2})/(\d{1,2})"),
    re.compile(r"^(\d{1,2})-(\d{1,2})-(\d{4})"),
]


def _to_iso_date(s: str) -> str | None:
    s = s.strip()
    if not s:
        return None
    for rx in _DATE_RX:
        m = rx.match(s)
        if not m:
            continue
        g = m.groups()
        if len(g[0]) == 4:
            y, mo, d = g
        else:
            mo, d, y = g
        try:
            return f"{int(y):04d}-{int(mo):02d}-{int(d):02d}"
        except ValueError:
            return None
    return None


# ─── NARA Catalog API ─────────────────────────────────────────────
# Public JSON API at catalog.archives.gov. No auth required for reads.
# Docs: https://catalog.archives.gov/api/v2/api-docs
def parse_nara_catalog(group_cfg: dict, limit: int | None, *, fetch_opts) -> tuple[list[Record], list[str]]:
    import os

    api_key = os.environ.get("NARA_API_KEY", "").strip()
    if not api_key:
        print(
            "[nara_catalog] NARA_API_KEY env var not set — skipping. "
            "Request a free key at Catalog_API@nara.gov and add it as a repo secret.",
            file=sys.stderr,
        )
        return [], []

    out: list[Record] = []
    seen_naids: set[str] = set()
    artifacts: list[str] = []

    base = "https://catalog.archives.gov/api/v2/records/search"
    headers = {"x-api-key": api_key, "Accept": "application/json"}
    for query in group_cfg.get("queries", []):
        q = query["q"]
        per_query_limit = int(query.get("limit", 200))
        # Page in 100-row chunks (API max).
        offset = 0
        page_size = 100
        while offset < per_query_limit:
            url = (
                f"{base}?q={requests.utils.quote(q)}"
                f"&limit={min(page_size, per_query_limit - offset)}"
                f"&offset={offset}"
            )
            artifacts.append(url)
            try:
                text = fetch(url, **fetch_opts(is_hub=True), extra_headers=headers)
                payload = json.loads(text)
            except Exception as e:
                print(f"  [nara_catalog] skip {q} @ {offset}: {e}", file=sys.stderr)
                break
            hits = (
                payload.get("body", {})
                .get("hits", {})
                .get("hits", [])
            )
            if not hits:
                break
            for hit in hits:
                src = hit.get("_source") or {}
                rec = src.get("record") or {}
                naid = str(rec.get("naId") or hit.get("_id") or "")
                if not naid or naid in seen_naids:
                    continue
                seen_naids.add(naid)
                title = (rec.get("title") or "").strip()
                if not title:
                    continue
                # Date heuristics
                date_iso = ""
                dates = rec.get("productionDates") or rec.get("inclusiveDates") or []
                if isinstance(dates, list) and dates:
                    first = dates[0]
                    if isinstance(first, dict):
                        date_iso = (
                            _to_iso_date(str(first.get("logicalDate", "")))
                            or _to_iso_date(str(first.get("startDate", "")))
                            or ""
                        )
                # Best-effort description
                scope = (rec.get("scopeAndContentNote") or "").strip()
                source_url = f"https://catalog.archives.gov/id/{naid}"
                out.append(
                    Record(
                        title=title,
                        agency=group_cfg["agency"],
                        unsealed_date=date_iso,
                        collection_id=group_cfg["collection_id_default"],
                        source_url=source_url,
                        description=scope[:500] if scope else None,
                        source_artifact_url=url,
                    )
                )
                if limit and len(out) >= limit:
                    return out, artifacts
            offset += page_size
    return out, artifacts


# ─── AARO parser ──────────────────────────────────────────────────
# AARO Cases-and-Reports page: each release sits inside a list-item that
# contains a date string like "October 2024" near the PDF anchor.
_MONTH_RX = re.compile(
    r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*\s+(\d{1,2},\s*)?(\d{4})\b",
    re.I,
)
_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
}


def _date_near(node) -> str:
    """Walk up the DOM looking for a Month-Year string."""
    cur = node
    for _ in range(4):
        if cur is None:
            break
        text = cur.get_text(" ", strip=True) if hasattr(cur, "get_text") else ""
        m = _MONTH_RX.search(text)
        if m:
            mo = _MONTHS[m.group(1).lower()[:4].rstrip()]
            day = int((m.group(2) or "1").strip(", ").strip()) if m.group(2) else 1
            y = int(m.group(3))
            return f"{y:04d}-{mo:02d}-{day:02d}"
        cur = getattr(cur, "parent", None)
    return ""


def parse_aaro(group_cfg: dict, limit: int | None, *, fetch_opts) -> tuple[list[Record], list[str]]:
    out: list[Record] = []
    artifacts: list[str] = []
    for hub in group_cfg["urls"]:
        artifacts.append(hub)
        html = fetch(hub, **fetch_opts(is_hub=True))
        soup = BeautifulSoup(html, "html.parser")
        for a in soup.find_all("a", href=True):
            href = urljoin(hub, a["href"])
            if not href.lower().endswith(".pdf"):
                continue
            title = (a.get_text(strip=True) or Path(href).stem).strip()
            date = _date_near(a)
            out.append(
                Record(
                    title=title or "(untitled AARO release)",
                    agency=group_cfg["agency"],
                    unsealed_date=date,
                    collection_id=group_cfg["collection_id_default"],
                    source_url=href,
                    source_artifact_url=hub,
                )
            )
            if limit and len(out) >= limit:
                return out, artifacts
    return out, artifacts


# ─── State FRUS parser ────────────────────────────────────────────
# Foreign Relations of the United States — the authoritative declassified
# diplomatic record. 551 EPUB volumes covering 1861-1989+, each containing
# 100-400 documents with clean structure: <h3> title, <p class="dateline">
# original document date, body paragraphs. We use the EPUBs (rather than
# crawling per-document URLs) because robots.txt has Crawl-delay: 20 — at
# 1 doc/20s the per-document path would take 116 days; one EPUB per
# volume × 551 volumes × 20s = ~3hr first run, then near-zero (FRUS
# publishes ~1-3 volumes/year).
_FRUS_EBOOKS_INDEX = "https://history.state.gov/historicaldocuments/ebooks"
_FRUS_EPUB_RX = re.compile(
    r"https://static\.history\.state\.gov/frus/([a-z0-9-]+)/ebook/[^.]+\.epub", re.I
)
_FRUS_TITLE_RX = re.compile(r"<h3[^>]*>(.*?)</h3>", re.S | re.I)
# Modern volumes: <p class="dateline">. Old volumes: <div class="opener">.
_FRUS_DATELINE_RX = re.compile(
    r'<p[^>]+class="dateline"[^>]*>(.*?)</p>', re.S | re.I
)
_FRUS_OPENER_RX = re.compile(
    r'<div[^>]+class="opener"[^>]*>(.*?)</div>', re.S | re.I
)
_FRUS_ANY_P_RX = re.compile(r"<p[^>]*>(.*?)</p>", re.S | re.I)
_FRUS_DOCNUM_RX = re.compile(r"^\s*\[document\s+\d+\]\s*$", re.I)
_FRUS_OPF_TITLE_RX = re.compile(r"<dc:title>(.*?)</dc:title>", re.S | re.I)
_FRUS_MONTH_DAY_YEAR_RX = re.compile(
    r"(?:^|,\s*)(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|"
    r"jul(?:y)?|aug(?:ust)?|sep(?:tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)"
    r"\s+(\d{1,2}),?\s+(\d{4})",
    re.I,
)


def _strip_html(s: str) -> str:
    # Drop footnote/superscript content entirely (FRUS titles have trailing
    # <sup>1</sup> markers we don't want bleeding into stored data).
    s = re.sub(r"<sup\b[^>]*>.*?</sup>", "", s, flags=re.S | re.I)
    s = re.sub(r"<[^>]+>", "", s)
    s = (
        s.replace("&amp;", "&")
        .replace("&lt;", "<")
        .replace("&gt;", ">")
        .replace("&quot;", '"')
        .replace("&#x2014;", "—")
        .replace("&#x2013;", "–")
        .replace("&#x2019;", "’")
        .replace("&nbsp;", " ")
    )
    return re.sub(r"\s+", " ", s).strip()


def _frus_parse_dateline(text: str) -> str:
    m = _FRUS_MONTH_DAY_YEAR_RX.search(text)
    if not m:
        # Try "Month YYYY" without day
        m2 = re.search(
            r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\s+(\d{4})\b",
            text, re.I,
        )
        if not m2:
            return ""
        mo = _MONTHS[m2.group(1).lower()[:3]]
        y = int(m2.group(2))
        return f"{y:04d}-{mo:02d}-01"
    mo = _MONTHS[m.group(1).lower()[:3]]
    day = int(m.group(2))
    y = int(m.group(3))
    return f"{y:04d}-{mo:02d}-{day:02d}"


_FRUS_PUB_DATE_FILE = ROOT / "ingest" / "frus_pub_dates.json"
_FRUS_TEI_URL = "https://raw.githubusercontent.com/HistoryAtState/frus/master/volumes/{}.xml"
_FRUS_PUB_DATE_RX = re.compile(
    r'<date[^>]+type="publication-date"[^>]*>([^<]+)</date>', re.I
)


def _frus_load_pub_dates() -> dict[str, str]:
    if _FRUS_PUB_DATE_FILE.exists():
        try:
            return json.loads(_FRUS_PUB_DATE_FILE.read_text("utf-8"))
        except Exception:
            return {}
    return {}


def _frus_save_pub_dates(d: dict[str, str]) -> None:
    _FRUS_PUB_DATE_FILE.write_text(json.dumps(d, indent=2, sort_keys=True), "utf-8")


def _frus_fetch_pub_date(vol_id: str, impersonate: str | None) -> str:
    """Fetch volume publication date by Range-GETting the first 8KB of the
    volume's TEI-XML from the HistoryAtState/frus GitHub mirror. No
    crawl-delay needed (GitHub raw on Fastly)."""
    url = _FRUS_TEI_URL.format(vol_id)
    try:
        if impersonate and _HAS_CFFI:
            r = _cffi_requests.get(
                url, impersonate=impersonate, timeout=TIMEOUT,
                headers={"Range": "bytes=0-8191"},
            )
        else:
            r = requests.get(
                url, timeout=TIMEOUT,
                headers={"user-agent": USER_AGENT, "From": FROM_HEADER,
                         "Range": "bytes=0-8191"},
            )
        if r.status_code not in (200, 206):
            return ""
        text = r.text
    except Exception:
        return ""
    m = _FRUS_PUB_DATE_RX.search(text)
    if not m:
        return ""
    raw = m.group(1).strip()
    # Normalize: "2006" → "2006-01-01"; "2006-04-15" → as-is.
    if re.fullmatch(r"\d{4}", raw):
        return f"{raw}-01-01"
    iso = _to_iso_date(raw)
    return iso or ""


def parse_frus(group_cfg: dict, limit: int | None, *, fetch_opts) -> tuple[list[Record], list[str]]:
    import zipfile
    import io as _io

    out: list[Record] = []
    artifacts: list[str] = []

    index_html = fetch(_FRUS_EBOOKS_INDEX, **fetch_opts(is_hub=True))
    epub_urls = sorted({m.group(0): m.group(1) for m in _FRUS_EPUB_RX.finditer(index_html)}.items())
    print(f"[frus] {len(epub_urls)} EPUB volumes discovered", file=sys.stderr)

    # Prefetch volume publication dates from the FRUS GitHub mirror. This is
    # cheap (~0.5s × 551 = ~5min) since GitHub raw is on Fastly with no
    # crawl-delay. Persisted in ingest/frus_pub_dates.json so we only refetch
    # for volumes we haven't seen.
    pub_dates = _frus_load_pub_dates()
    impersonate = group_cfg.get("impersonate")
    needed = [vid for _u, vid in epub_urls if vid not in pub_dates]
    if needed:
        print(f"[frus] fetching publication dates for {len(needed)} new volumes...", file=sys.stderr)
        for i, vid in enumerate(needed):
            d = _frus_fetch_pub_date(vid, impersonate)
            pub_dates[vid] = d
            if (i + 1) % 50 == 0:
                _frus_save_pub_dates(pub_dates)
                print(f"  [frus] pub-dates {i + 1}/{len(needed)}", file=sys.stderr)
        _frus_save_pub_dates(pub_dates)

    for epub_url, vol_id in epub_urls:
        artifacts.append(epub_url)
        try:
            data = fetch(epub_url, binary=True, **fetch_opts(is_hub=False))
        except Exception as e:
            print(f"  [frus] skip {vol_id}: {e}", file=sys.stderr)
            continue
        try:
            z = zipfile.ZipFile(_io.BytesIO(data))
        except zipfile.BadZipFile as e:
            print(f"  [frus] bad zip {vol_id}: {e}", file=sys.stderr)
            continue

        # Volume metadata from the OPF — we use only the title. The
        # <dc:identifier> embeds the EPUB *rebuild* date (often 2018+ even
        # for an 1861 volume), so it's a poor proxy for when the volume
        # was actually published / declassified. We use the document's own
        # date as `unsealed_date` instead (correct within months for old
        # volumes, off by years for modern volumes — but always extractable
        # and consistent with how users browse FRUS).
        opf_name = next((n for n in z.namelist() if n.endswith(".opf")), None)
        vol_title = vol_id
        if opf_name:
            opf = z.read(opf_name).decode("utf-8", "replace")
            mt = _FRUS_OPF_TITLE_RX.search(opf)
            if mt:
                vol_title = _strip_html(mt.group(1))

        # Per-document files: OEBPS/dN.html (numbered docs only — skip
        # frontmatter like cover/title/preface).
        doc_files = sorted(
            [n for n in z.namelist() if re.search(r"/d\d+\.html$", n)],
            key=lambda n: int(re.search(r"/d(\d+)\.html$", n).group(1)),
        )
        for doc_name in doc_files:
            doc_num = re.search(r"/d(\d+)\.html$", doc_name).group(1)
            body = z.read(doc_name).decode("utf-8", "replace")

            mt = _FRUS_TITLE_RX.search(body)
            title = _strip_html(mt.group(1)) if mt else ""
            # Strip leading "N. " numbering
            title = re.sub(r"^\d+\.\s+", "", title)
            if not title:
                continue

            # Date: dateline (modern) → opener div (old) → fallback scan.
            md = _FRUS_DATELINE_RX.search(body)
            if md:
                doc_date_text = _strip_html(md.group(1))
            else:
                mo = _FRUS_OPENER_RX.search(body)
                doc_date_text = _strip_html(mo.group(1)) if mo else ""
            doc_date = _frus_parse_dateline(doc_date_text) if doc_date_text else ""

            # First substantive paragraph (skip "[Document N]" markers).
            preview = None
            for pm in _FRUS_ANY_P_RX.finditer(body):
                p = _strip_html(pm.group(1))
                if not p or _FRUS_DOCNUM_RX.match(p):
                    continue
                # Skip the dateline paragraph itself (we already have it).
                if p == doc_date_text:
                    continue
                preview = p[:400]
                break

            description = None
            if doc_date_text or preview:
                description = " // ".join(
                    [s for s in (doc_date_text, preview) if s]
                )[:500]

            vol_pub_date = pub_dates.get(vol_id, "")
            out.append(
                Record(
                    title=title,
                    agency=group_cfg["agency"],
                    # unsealed_date = volume publication date (when this
                    # document became public). document_date = doc's own
                    # creation date. Fall back to doc_date for unsealed_date
                    # if the GitHub TEI mirror didn't have a pub date.
                    unsealed_date=vol_pub_date or doc_date,
                    document_date=doc_date or None,
                    collection_id=vol_title,
                    source_url=f"https://history.state.gov/historicaldocuments/{vol_id}/d{doc_num}",
                    description=description,
                    source_artifact_url=epub_url,
                )
            )
            if limit and len(out) >= limit:
                return out, artifacts
    return out, artifacts


PARSERS = {
    "nara_ndc": parse_nara_ndc,
    "nara_catalog": parse_nara_catalog,
    "aaro": parse_aaro,
    "frus": parse_frus,
    # dow_uap intentionally absent until a verified source URL exists.
}


# ─── SQL emission ─────────────────────────────────────────────────
def emit_sql(group: str, records: list[Record], seen_hashes: set[str] | None = None) -> Path:
    out_path = OUT_DIR / f"ingest_{group}.sql"
    # Filter out records whose content_hash is already in D1. Those would
    # INSERT OR IGNORE to no-ops, but each one still counts against the
    # 100k/day write quota on D1's Workers Free tier.
    seen_hashes = seen_hashes or set()
    fresh = [r for r in records if r.content_hash not in seen_hashes]
    skipped = len(records) - len(fresh)
    with out_path.open("w", encoding="utf-8") as f:
        f.write(
            f"-- UNSEALED ingest — {group} — {len(fresh)} new records "
            f"({skipped} already in D1, skipped) — run {RUN_ID}\n"
        )
        for r in fresh:
            f.write(
                "INSERT OR IGNORE INTO records "
                "(title, agency, unsealed_date, collection_id, source_url, "
                "description, thumbnail_url, source_artifact_url, is_sealed, "
                "document_date, ingest_run_id, content_hash) "
                "VALUES ("
                f"{_q(r.title)}, {_q(r.agency)}, {_q(r.unsealed_date)}, {_q(r.collection_id)}, "
                f"{_q(r.source_url)}, {_q(r.description)}, {_q(r.thumbnail_url)}, "
                f"{_q(r.source_artifact_url)}, {int(r.is_sealed)}, "
                f"{_q(r.document_date)}, {_q(RUN_ID)}, {_q(r.content_hash)}"
                ");\n"
            )
    return out_path


def _q(v: str | None) -> str:
    if v is None or v == "":
        return "NULL"
    return "'" + str(v).replace("'", "''") + "'"


# ─── CLI ──────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--group", help="run only one source group")
    ap.add_argument("--dry-run", action="store_true")
    ap.add_argument("--limit", type=int, default=None)
    ap.add_argument("--no-cache", action="store_true", help="bypass on-disk cache")
    ap.add_argument(
        "--cache-ttl", type=int, default=None,
        help="hub-page cache TTL in seconds (default: 7d). 0 = always re-fetch."
    )
    ap.add_argument(
        "--self-test", action="store_true",
        help="run internal asserts (date parsing) and exit"
    )
    args = ap.parse_args()

    if args.self_test:
        _self_test()
        print("self-test ok", file=sys.stderr)
        return 0

    cfg = json.loads(SOURCES_FILE.read_text("utf-8"))
    groups = [args.group] if args.group else [g for g in cfg if not g.startswith("_")]

    seen_artifacts = load_seen_artifacts()
    seen_hashes = load_seen_hashes()
    if seen_hashes:
        print(f"loaded {len(seen_hashes)} seen content_hashes (skip-already-in-D1 mode)", file=sys.stderr)
    report: dict = {
        "run_id": RUN_ID,
        "groups": {},
        "total_records": 0,
        "new_artifacts": {},
    }

    def make_fetch_opts(impersonate: str | None, crawl_delay_sec: int):
        def fetch_opts(*, is_hub: bool) -> dict:
            ttl = (
                args.cache_ttl
                if args.cache_ttl is not None
                else (DEFAULT_HUB_TTL_SEC if is_hub else DEFAULT_ARTIFACT_TTL_SEC)
            )
            opts = {"no_cache": args.no_cache, "ttl_sec": ttl}
            if impersonate:
                opts["impersonate"] = impersonate
            if crawl_delay_sec:
                opts["crawl_delay_sec"] = crawl_delay_sec
            return opts
        return fetch_opts

    for g in groups:
        if g not in cfg:
            print(f"unknown group: {g}", file=sys.stderr)
            continue
        if g not in PARSERS:
            print(f"[{g}] no parser registered — skipping", file=sys.stderr)
            report["groups"][g] = {"status": "skipped_no_parser"}
            continue
        if not (cfg[g].get("urls") or cfg[g].get("queries")):
            print(f"[{g}] no urls/queries configured — skipping", file=sys.stderr)
            report["groups"][g] = {"status": "skipped_no_urls"}
            continue
        print(f"[{g}] parsing...", file=sys.stderr)
        try:
            recs, artifacts = PARSERS[g](cfg[g], args.limit, fetch_opts=make_fetch_opts(cfg[g].get("impersonate"), int(cfg[g].get("crawl_delay_sec", 0))))
        except Exception as e:
            print(f"[{g}] FAILED: {e}", file=sys.stderr)
            report["groups"][g] = {"status": "error", "error": str(e)}
            continue
        prev = set(seen_artifacts.get(g, []))
        new = sorted(set(artifacts) - prev)
        seen_artifacts[g] = sorted(set(artifacts) | prev)
        if new:
            print(f"[{g}] {len(new)} NEW artifacts since last run", file=sys.stderr)
            report["new_artifacts"][g] = new

        # Filter against seen_hashes so the report + emitted SQL reflect
        # only records that aren't already in D1.
        fresh_recs = [r for r in recs if r.content_hash not in seen_hashes]
        already_in_d1 = len(recs) - len(fresh_recs)
        print(
            f"[{g}] {len(recs)} parsed, {len(fresh_recs)} new ({already_in_d1} already in D1), "
            f"{len(artifacts)} artifacts",
            file=sys.stderr,
        )
        report["groups"][g] = {
            "status": "ok",
            "records": len(fresh_recs),
            "records_parsed": len(recs),
            "records_already_in_d1": already_in_d1,
            "artifacts": len(artifacts),
            "new_artifacts": len(new),
        }
        report["total_records"] += len(fresh_recs)
        if args.dry_run:
            for r in fresh_recs[:5]:
                print(json.dumps(asdict(r), indent=2))
            continue
        path = emit_sql(g, recs, seen_hashes=seen_hashes)
        print(f"[{g}] wrote {path.relative_to(ROOT)}", file=sys.stderr)

    if not args.dry_run:
        save_seen_artifacts(seen_artifacts)
        report_path = OUT_DIR / "ingest_report.json"
        report_path.write_text(json.dumps(report, indent=2), "utf-8")
        print(f"wrote {report_path.relative_to(ROOT)}", file=sys.stderr)

    print(f"total: {report['total_records']} records", file=sys.stderr)
    return 0


def _self_test() -> None:
    cases = [
        # FY-compact: FY2019-Q2 → fiscal Q2 = calendar Q1 of 2019
        ("https://declassification.blogs.archives.gov/wp-content/uploads/sites/16/2019/07/FY2019-Q2-Release-List-Excel-Format.xlsx", "2019-03-31"),
        ("https://declassification.blogs.archives.gov/wp-content/uploads/sites/16/2019/08/FY2019-Q3-Release-List-Excel-Format.xlsx", "2019-06-30"),
        ("https://declassification.blogs.archives.gov/wp-content/uploads/sites/16/2019/10/FY2019-Q4-Release-List-Excel-Format.xlsx", "2019-09-30"),
        ("https://declassification.blogs.archives.gov/wp-content/uploads/sites/16/2020/01/FY2020-Q1-Release-List-Excel-Format.xlsx", "2019-12-31"),
        # Calendar-year ordinal quarter
        ("https://declassification.blogs.archives.gov/wp-content/uploads/sites/16/2023/10/2023-3rd-quarter-release-list.xlsx", "2023-09-30"),
        ("https://declassification.blogs.archives.gov/wp-content/uploads/sites/16/2023/10/2023-4th-Quarter-Release-List-October-6th.xlsx", "2023-12-31"),
        # Existing patterns still work
        ("https://www.archives.gov/files/declassification/ndc/2024-ndc-1st-quarter-release-list-excel.xlsx", "2024-03-31"),
        ("https://www.archives.gov/files/2nd-quarter-release-list-fy-26.xlsx", "2026-03-31"),
        # released-entries-MM-YY (legacy monthly NDC lists) → month-end of 20YY-MM
        ("https://www.archives.gov/declassification/ndc/reports/released-entries-05-12.xls", "2012-05-31"),
        ("https://www.archives.gov/declassification/ndc/reports/released-entries-07-12.xls", "2012-07-31"),
        ("https://www.archives.gov/declassification/ndc/reports/released-entries-04-13.xls", "2013-04-30"),
        # q4-2022 form (quarter then calendar year)
        ("https://www.archives.gov/files/declassification/ndc/release-list-q4-2022-excel.xlsx", "2022-12-31"),
        # 2nd-qt-2023 form (ordinal + qt/qtr abbreviation)
        ("https://www.archives.gov/files/declassification/ndc/reports/release-list-projects-for-2nd-qt-2023.xlsx", "2023-06-30"),
        # q1-february-23 form (quarter, month-word, 2-digit year)
        ("https://www.archives.gov/files/declassification/release-list-q1-february-23.xlsx", "2023-03-31"),
    ]
    for url, expected in cases:
        got = _date_from_filename(url)
        assert got == expected, f"{url}\n  expected {expected!r}, got {got!r}"


if __name__ == "__main__":
    sys.exit(main())
